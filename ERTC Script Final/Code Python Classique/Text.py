from openpyxl import load_workbook

excel_file = r'C:\Users\malak\Downloads\ERTC Script Finale\ERTC Script Final\Code Python Classique\Export_Excel\TABLEAU GENERAL MODELE V2.xlsx'

#Lire le fichier Excel dans un DataFrame
#df = pd.read_excel(excel_file, engine='openpyxl')
wb = load_workbook(excel_file)

#selectionner la premiére feuille
ws = wb.active

#ouvrir le fichier text pour écrire les résultats
with open("output.txt", "w", newline='',encoding = "utf-8") as file:
    for column in ws.iter_cols():
        # Récupérer le contenu de la première cellule de la colonne (ligne 1)
        line1_content = column[0].value
        # Récupérer le contenu de la deuxième cellule de la colonne (ligne 2)
        line2_content = column[1].value
        # Écrire le résultat dans le fichier texte
        file.write(f"{line1_content} : {line2_content}\n")

print("Fichier texte généré avec succès !")





